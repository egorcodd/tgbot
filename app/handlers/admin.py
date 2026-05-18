import asyncio
import logging
from datetime import datetime
from io import BytesIO

from aiogram import Bot, F, Router
from aiogram.exceptions import TelegramForbiddenError, TelegramRetryAfter, TelegramBadRequest
from aiogram.filters import Command
from aiogram.types import BufferedInputFile, Message
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sqlalchemy import func, select
from sqlalchemy.orm import selectinload

from app.config import settings
from app.db.database import SessionLocal
from app.db import repo
from app.db.models import Answer, Broadcast, Lead
from app.quiz.questions import QUESTIONS

router = Router()
log = logging.getLogger(__name__)

# Telegram allows ~30 msg/s to different users; stay below to be safe.
BROADCAST_RATE_PER_SEC = 25


def _is_admin(user_id: int) -> bool:
    return user_id in settings.admin_id_list


@router.message(Command("getfileid"))
async def cmd_getfileid(message: Message) -> None:
    if message.from_user is None or not _is_admin(message.from_user.id):
        return
    target = message.reply_to_message or message
    bits: list[str] = []
    if target.video_note:
        bits.append(f"video_note: <code>{target.video_note.file_id}</code>")
    if target.video:
        bits.append(f"video: <code>{target.video.file_id}</code>")
    if target.photo:
        bits.append(f"photo: <code>{target.photo[-1].file_id}</code>")
    if target.document:
        bits.append(f"document: <code>{target.document.file_id}</code>")
    if target.animation:
        bits.append(f"animation: <code>{target.animation.file_id}</code>")
    if not bits:
        await message.answer("Пришли (или сделай reply на) видео-кружок / видео / фото / документ — отдам file_id.")
        return
    await message.answer("\n".join(bits), parse_mode="HTML")


@router.message(Command("stats"))
async def cmd_stats(message: Message) -> None:
    if message.from_user is None or not _is_admin(message.from_user.id):
        return
    async with SessionLocal() as session:
        total = await session.scalar(select(func.count(Lead.id))) or 0
        finished = await session.scalar(select(func.count(Lead.id)).where(Lead.finished.is_(True))) or 0
        active = await session.scalar(select(func.count(Lead.id)).where(Lead.is_active.is_(True))) or 0
        # drop-off per step
        rows = await session.execute(
            select(Lead.current_step, func.count(Lead.id)).group_by(Lead.current_step).order_by(Lead.current_step)
        )
        by_step = list(rows.all())

    lines = [
        f"<b>Лиды:</b> {total}",
        f"<b>Активные:</b> {active}",
        f"<b>Прошли квиз:</b> {finished}",
        "",
        "<b>Дроп по шагам (current_step → людей):</b>",
    ]
    for step, n in by_step:
        lines.append(f"  {step}: {n}")
    await message.answer("\n".join(lines), parse_mode="HTML")


@router.message(Command("export"))
async def cmd_export(message: Message) -> None:
    """Выгружает всех лидов и их ответы в .xlsx и присылает файлом сюда же."""
    if message.from_user is None or not _is_admin(message.from_user.id):
        return

    async with SessionLocal() as session:
        result = await session.execute(
            select(Lead).options(selectinload(Lead.answers)).order_by(Lead.created_at)
        )
        leads = result.scalars().all()

    wb = Workbook()
    ws = wb.active
    ws.title = "leads"

    question_keys = [q["key"] for q in QUESTIONS]
    question_texts = {q["key"]: q["text"] for q in QUESTIONS}

    base_headers = [
        "tg_id", "username", "имя",
        "источник (utm)", "создан", "прошёл квиз", "завершил",
        "активен", "шаг",
    ]
    headers = base_headers + [question_texts[k] for k in question_keys]
    ws.append(headers)

    for lead in leads:
        answers_map = {a.question_key: a.option_text for a in lead.answers}
        full_name = " ".join(p for p in (lead.first_name, lead.last_name) if p) or ""
        row = [
            lead.tg_id,
            lead.username or "",
            full_name,
            lead.source or "",
            lead.created_at.strftime("%Y-%m-%d %H:%M") if lead.created_at else "",
            "да" if lead.finished else "нет",
            lead.finished_at.strftime("%Y-%m-%d %H:%M") if lead.finished_at else "",
            "да" if lead.is_active else "нет",
            lead.current_step,
        ]
        row += [answers_map.get(k, "") for k in question_keys]
        ws.append(row)

    # авто-ширина колонок (грубая)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for cell in ws[get_column_letter(col_idx)]:
            v = cell.value
            if v is not None:
                max_len = max(max_len, min(len(str(v)), 60))
        ws.column_dimensions[get_column_letter(col_idx)].width = max_len + 2

    # шапка жирная
    for cell in ws[1]:
        cell.font = cell.font.copy(bold=True)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"leads_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await message.answer_document(
        document=BufferedInputFile(buf.read(), filename=filename),
        caption=f"Лидов: {len(leads)}",
    )


@router.message(Command("broadcast"))
async def cmd_broadcast(message: Message, bot: Bot) -> None:
    if message.from_user is None or not _is_admin(message.from_user.id):
        return

    src = message.reply_to_message
    if src is None:
        await message.answer(
            "Сделай <b>reply</b> на сообщение (текст / фото / видео / кружок / документ) "
            "и снова отправь /broadcast.\n\n"
            "Опционально: <code>/broadcast finished</code> — только тем, кто прошёл квиз.",
            parse_mode="HTML",
        )
        return

    parts = (message.text or "").split(maxsplit=1)
    segment = parts[1].strip().lower() if len(parts) > 1 else "all"

    async with SessionLocal() as session:
        q = select(Lead.tg_id).where(Lead.is_active.is_(True))
        if segment == "finished":
            q = q.where(Lead.finished.is_(True))
        elif segment == "unfinished":
            q = q.where(Lead.finished.is_(False))
        elif segment.startswith("answer:"):
            # e.g. "answer:studied=python" — segment by quiz answer
            try:
                _, kv = segment.split(":", 1)
                qkey, okey = kv.split("=", 1)
            except ValueError:
                await message.answer("Формат: /broadcast answer:question_key=option_key")
                return
            q = q.where(Lead.id.in_(
                select(Answer.lead_id).where(Answer.question_key == qkey, Answer.option_key == okey)
            ))
        elif segment != "all":
            await message.answer(
                "Поддерживаемые сегменты: <code>all</code>, <code>finished</code>, "
                "<code>unfinished</code>, <code>answer:KEY=OPT</code>",
                parse_mode="HTML",
            )
            return

        recipients = [tg_id for (tg_id,) in (await session.execute(q)).all()]

    if not recipients:
        await message.answer("Под этот сегмент никто не подходит.")
        return

    await message.answer(f"Рассылка стартует. Получателей: <b>{len(recipients)}</b>", parse_mode="HTML")

    sent = 0
    failed = 0
    started = asyncio.get_event_loop().time()
    for i, tg_id in enumerate(recipients):
        try:
            await src.copy_to(chat_id=tg_id)
            sent += 1
        except TelegramRetryAfter as e:
            await asyncio.sleep(e.retry_after + 1)
            try:
                await src.copy_to(chat_id=tg_id)
                sent += 1
            except Exception as exc:
                log.warning("retry failed for %s: %s", tg_id, exc)
                failed += 1
        except TelegramForbiddenError:
            # user blocked the bot — flag inactive
            async with SessionLocal() as session:
                await repo.mark_blocked(session, tg_id)
            failed += 1
        except TelegramBadRequest as exc:
            log.warning("broadcast bad request to %s: %s", tg_id, exc)
            failed += 1
        except Exception as exc:
            log.warning("broadcast unexpected error to %s: %s", tg_id, exc)
            failed += 1

        # rate-limit
        if (i + 1) % BROADCAST_RATE_PER_SEC == 0:
            elapsed = asyncio.get_event_loop().time() - started
            target = (i + 1) / BROADCAST_RATE_PER_SEC
            if elapsed < target:
                await asyncio.sleep(target - elapsed)

    async with SessionLocal() as session:
        session.add(Broadcast(
            admin_tg_id=message.from_user.id,
            kind=_detect_kind(src),
            payload=(src.text or src.caption or "[media]")[:1000],
            caption=src.caption,
            segment=segment,
            total=len(recipients),
            sent=sent,
            failed=failed,
        ))
        await session.commit()

    await message.answer(
        f"Готово. Отправлено: <b>{sent}</b>, не доставлено: <b>{failed}</b>",
        parse_mode="HTML",
    )


def _detect_kind(msg: Message) -> str:
    if msg.video_note:
        return "video_note"
    if msg.video:
        return "video"
    if msg.photo:
        return "photo"
    if msg.document:
        return "document"
    if msg.animation:
        return "animation"
    return "text"
